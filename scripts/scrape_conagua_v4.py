#!/usr/bin/env python3
"""
scrape_conagua_v4.py - fixes especificos basados en inspeccion real
=====================================================================

CAMBIOS RESPECTO v3:
  - clima_municipal:  Usa Open-Meteo (gratis, sin auth) para 32 capitales estatales.
                      SMN no expone tabla de temperaturas por estado.
  - alertas_meteo:    Parsea el boletin de texto del SMN como UNA alerta nacional
                      (la URL /es/avisos-tiempo da 404).
  - monitor_sequia:   Parser correcto del XLSX wide-format: 2479 municipios x 40
                      columnas de fecha. Agrega por estado calculando % en cada D0-D4.
  - presas_cuencas:   SINA JSON con timeout largo y verify=False.

FIX 2026-06-03 (anti-zombie):
  - Eliminado supa_truncate() en las 4 tablas. Reemplazado por UPSERT con
    on_conflict en llave natural (fecha_X, estado/presa/muni). Si la peticion
    falla, la tabla NUNCA queda vacia. Conservamos historial automaticamente.

Variables de entorno:
  SUPABASE_URL
  SUPABASE_SERVICE_ROLE_KEY
"""

import os
import re
import io
import datetime as dt
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# FIX-27 2026-05-21: cloudscraper para bypass de anti-bot Imperva en gob.mx
# FIX-27 fix 3: NoVerifyAdapter para SMN cert chain broken
import ssl as _ssl
from requests.adapters import HTTPAdapter as _HA
try:
    from urllib3.util.ssl_ import create_urllib3_context as _ctx
except ImportError:
    _ctx = None

class _NoVerify(_HA):
    def init_poolmanager(self, *a, **kw):
        if _ctx:
            c = _ctx()
            c.check_hostname = False
            c.verify_mode = _ssl.CERT_NONE
            kw['ssl_context'] = c
        return super().init_poolmanager(*a, **kw)

try:
    import cloudscraper
    _CS = cloudscraper.create_scraper(browser={'browser': 'chrome', 'platform': 'darwin', 'mobile': False})
    _CS.verify = False
    _CS.mount('https://', _NoVerify())
    def gob_get(url, **kw):
        kw.setdefault('verify', False)
        return _CS.get(url, **kw)
    print("  [cloudscraper] OK -- anti-bot gob.mx + NoVerifyAdapter activo")
except ImportError:
    _S = requests.Session()
    _S.mount('https://', _NoVerify())
    _S.verify = False
    def gob_get(url, **kw):
        kw.setdefault('verify', False)
        return _S.get(url, **kw)
    print("  [cloudscraper] NO disponible -- requests con NoVerifyAdapter")

SUPABASE_URL = os.environ['SUPABASE_URL'].rstrip('/')
SERVICE_KEY  = os.environ['SUPABASE_SERVICE_ROLE_KEY']

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (compatible; TequioCivicBot/4.0; +https://tequio.app)',
    'Accept': 'text/html,application/xhtml+xml,application/json,*/*',
    'Accept-Language': 'es-MX,es;q=0.9,en;q=0.8',
}
TIMEOUT = 90


# =====================================================================
# HELPERS
# =====================================================================
def supa_post(table, rows, on_conflict=None):
    if not rows: return 0
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if on_conflict:
        url += f"?on_conflict={on_conflict}"
        prefer = 'resolution=merge-duplicates,return=minimal'
    else:
        prefer = 'return=minimal'
    try:
        r = requests.post(url, json=rows, headers={
            'Content-Type': 'application/json',
            'apikey': SERVICE_KEY,
            'Authorization': f'Bearer {SERVICE_KEY}',
            'Prefer': prefer,
        }, timeout=TIMEOUT)
        if not r.ok:
            print(f"   [WARN] Supabase {table} {r.status_code}: {r.text[:250]}")
            return 0
        return len(rows)
    except Exception as e:
        print(f"   [WARN] Supabase {table} excepcion: {e}")
        return 0


def _num(v):
    if v is None: return None
    s = str(v).replace(',', '').strip()
    if s in ('', '-', 'N/D', 'NA', 'ND', 'S/D'): return None
    try: return float(s)
    except: return None


# =====================================================================
# 1) CLIMA MUNICIPAL - Open-Meteo para 32 capitales estatales
# =====================================================================
CAPITALES_MX = [
    ('Aguascalientes',     'Aguascalientes',       21.8853, -102.2916),
    ('Baja California',    'Mexicali',             32.6245, -115.4523),
    ('Baja California Sur','La Paz',               24.1426, -110.3128),
    ('Campeche',           'Campeche',             19.8301, -90.5349),
    ('Coahuila',           'Saltillo',             25.4232, -100.9979),
    ('Colima',             'Colima',               19.2452, -103.7241),
    ('Chiapas',            'Tuxtla Gutierrez',     16.7530, -93.1156),
    ('Chihuahua',          'Chihuahua',            28.6353, -106.0889),
    ('Ciudad de Mexico',   'Ciudad de Mexico',     19.4326, -99.1332),
    ('Durango',            'Durango',              24.0277, -104.6532),
    ('Guanajuato',         'Guanajuato',           21.0190, -101.2574),
    ('Guerrero',           'Chilpancingo',         17.5506, -99.5006),
    ('Hidalgo',            'Pachuca',              20.1011, -98.7591),
    ('Jalisco',            'Guadalajara',          20.6597, -103.3496),
    ('Mexico',             'Toluca',               19.2826, -99.6557),
    ('Michoacan',          'Morelia',              19.7008, -101.1844),
    ('Morelos',            'Cuernavaca',           18.9242, -99.2216),
    ('Nayarit',            'Tepic',                21.5039, -104.8946),
    ('Nuevo Leon',         'Monterrey',            25.6866, -100.3161),
    ('Oaxaca',             'Oaxaca de Juarez',     17.0732, -96.7266),
    ('Puebla',             'Puebla',               19.0414, -98.2063),
    ('Queretaro',          'Queretaro',            20.5888, -100.3899),
    ('Quintana Roo',       'Chetumal',             18.5036, -88.3055),
    ('San Luis Potosi',    'San Luis Potosi',      22.1565, -100.9855),
    ('Sinaloa',            'Culiacan',             24.8091, -107.3940),
    ('Sonora',             'Hermosillo',           29.0729, -110.9559),
    ('Tabasco',            'Villahermosa',         17.9892, -92.9475),
    ('Tamaulipas',         'Ciudad Victoria',      23.7369, -99.1411),
    ('Tlaxcala',           'Tlaxcala',             19.3139, -98.2400),
    ('Veracruz',           'Xalapa',               19.5438, -96.9102),
    ('Yucatan',            'Merida',               20.9674, -89.5926),
    ('Zacatecas',          'Zacatecas',            22.7709, -102.5832),
]


def scrape_clima_municipal():
    print("\n[CLIMA] Open-Meteo para 32 capitales estatales...")
    rows = []
    hoy = dt.date.today().isoformat()
    base = "https://api.open-meteo.com/v1/forecast"

    for estado, capital, lat, lon in CAPITALES_MX:
        try:
            params = {
                'latitude':  lat,
                'longitude': lon,
                'daily':     'temperature_2m_max,temperature_2m_min,precipitation_probability_max,weathercode,windspeed_10m_max,winddirection_10m_dominant',
                'timezone':  'America/Mexico_City',
                'forecast_days': 1,
            }
            r = None
            for intento in range(2):
                try:
                    r = requests.get(base, params=params, headers=HEADERS, timeout=60)
                    if r.ok:
                        break
                except Exception:
                    if intento == 0:
                        import time; time.sleep(2)
                        continue
                    r = None
            if not r or not r.ok:
                print(f"   [WARN] {capital}: HTTP {r.status_code if r else 'TIMEOUT'} tras 2 intentos")
                continue
            d = r.json().get('daily', {})
            if not d.get('temperature_2m_max'):
                continue
            wcode = int(d['weathercode'][0]) if d.get('weathercode') else 0
            desc_map = {
                0: 'Despejado', 1: 'Mayormente despejado', 2: 'Parcialmente nublado', 3: 'Nublado',
                45: 'Niebla', 48: 'Niebla con escarcha',
                51: 'Llovizna ligera', 53: 'Llovizna moderada', 55: 'Llovizna intensa',
                61: 'Lluvia ligera', 63: 'Lluvia moderada', 65: 'Lluvia intensa',
                71: 'Nieve ligera', 73: 'Nieve moderada', 75: 'Nieve intensa',
                80: 'Chubascos', 81: 'Chubascos fuertes', 82: 'Chubascos muy fuertes',
                95: 'Tormenta electrica', 96: 'Tormenta con granizo', 99: 'Tormenta severa',
            }
            estado_norm = estado.upper().replace(' ', '_')[:30]
            rows.append({
                'fecha_pronostico':  hoy,
                'municipio_id':      'capital_' + estado_norm,
                'municipio':         capital,
                'estado':            estado,
                'temp_max':          _num(d['temperature_2m_max'][0]),
                'temp_min':          _num(d['temperature_2m_min'][0]),
                'prob_lluvia':       int(d.get('precipitation_probability_max', [0])[0] or 0),
                'desc_cielo':        desc_map.get(wcode, f'WMO {wcode}'),
                'velocidad_viento':  _num(d.get('windspeed_10m_max', [None])[0]),
                'direccion_viento':  str(d.get('winddirection_10m_dominant', [''])[0] or ''),
                'fuente':            'Open-Meteo',
            })
        except Exception as e:
            print(f"   [WARN] {capital}: {e}")

    if not rows:
        print("   [ERR] Open-Meteo no devolvio nada.")
        return
    # FIX anti-zombie: UPSERT con on_conflict, sin truncate
    n = supa_post('clima_municipal', rows, on_conflict='fecha_pronostico,municipio_id')
    print(f"   [OK] {n} capitales con pronostico hoy (upsert)")


# =====================================================================
# 2) ALERTAS - boletin de texto SMN como alerta nacional
# =====================================================================
def scrape_alertas_meteo():
    print("\n[ALERTAS] Boletin meteorologico SMN como alerta nacional...")
    url = "https://smn.conagua.gob.mx/es/pronostico-meteorologico-general"
    try:
        r = requests.get(url, headers=HEADERS, timeout=TIMEOUT, verify=False)
        if not r.ok:
            print(f"   [WARN] HTTP {r.status_code}")
            return
        soup = BeautifulSoup(r.text, 'lxml')
        full = soup.get_text(' ', strip=True)
        m = re.search(r'(No\.\s*Aviso[:\s]*\d+.*?)(?=Pron[oó]stico|Para la|Servicios|$)', full, re.IGNORECASE | re.DOTALL)
        cuerpo = (m.group(1) if m else full[:3000])[:2000]
        m_titulo = re.search(r'([A-ZÁÉÍÓÚÑ\s,]{30,200})', cuerpo)
        titulo = m_titulo.group(1).strip()[:200] if m_titulo else 'Aviso SMN diario'
        low = cuerpo.lower()
        nivel = 'verde'
        if any(k in low for k in ['huracan', 'extremo', 'categoria 4', 'categoria 5']): nivel='rojo'
        elif any(k in low for k in ['intenso', 'severo', 'muy fuerte', 'onda de calor']): nivel='naranja'
        elif any(k in low for k in ['fuerte', 'moderado', 'lluvia', 'viento']):           nivel='amarillo'

        estados_mx = ['Aguascalientes','Baja California','Baja California Sur','Campeche','Chiapas',
                      'Chihuahua','Coahuila','Colima','Durango','Guanajuato','Guerrero','Hidalgo',
                      'Jalisco','Estado de México','Michoacán','Morelos','Nayarit','Nuevo León',
                      'Oaxaca','Puebla','Querétaro','Quintana Roo','San Luis Potosí','Sinaloa',
                      'Sonora','Tabasco','Tamaulipas','Tlaxcala','Veracruz','Yucatán','Zacatecas',
                      'Ciudad de México','CDMX']
        zona = ', '.join(e for e in estados_mx if e in cuerpo)[:300]

        ahora = dt.datetime.utcnow().isoformat()
        row = {
            'tipo':           'boletin_smn',
            'nombre':         titulo,
            'nivel':          nivel,
            'zona_afectada':  zona,
            'descripcion':    cuerpo,
            'vigente_desde':  ahora,
            'vigente_hasta':  None,
            'fuente':         'CONAGUA SMN',
            'url_oficial':    url,
        }
        # FIX anti-zombie: INSERT acumulativo (sin truncate). Tabla pequena, mantiene historico.
        # El UI debe pedir ORDER BY vigente_desde DESC LIMIT 1 para el aviso mas reciente.
        n = supa_post('alertas_meteo', [row])
        print(f"   [OK] {n} aviso nacional guardado (nivel={nivel})")
    except Exception as e:
        print(f"   [ERR] {e}")


# =====================================================================
# 3) MONITOR DE SEQUIA - XLSX wide-format real
# =====================================================================
def scrape_monitor_sequia():
    """Agrego por estado para la fecha mas reciente."""
    print("\n[SEQUIA] Monitor de Sequia (XLSX wide-format)...")
    url_xlsx = "https://smn.conagua.gob.mx/tools/RESOURCES/Monitor%20de%20Sequia%20en%20Mexico/MunicipiosSequia.xlsx"
    try:
        r = gob_get(url_xlsx, headers=HEADERS, timeout=TIMEOUT)
        if not r.ok:
            print(f"   [WARN] HTTP {r.status_code}")
            return
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("   [WARN] openpyxl no instalado.")
            return
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        ws = wb['MUNICIPIOS'] if 'MUNICIPIOS' in wb.sheetnames else wb.active

        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 10:
            print("   [WARN] Excel vacio.")
            return
        header = all_rows[0]
        FIRST_DATE_COL = 9
        last_col = FIRST_DATE_COL
        for col_idx in range(len(header)-1, FIRST_DATE_COL-1, -1):
            non_empty = sum(1 for row in all_rows[1:] if col_idx < len(row) and row[col_idx])
            if non_empty > 100:
                last_col = col_idx
                break
        excel_serial = header[last_col]
        if isinstance(excel_serial, (int, float)):
            fecha_corte = (dt.date(1899, 12, 30) + dt.timedelta(days=int(excel_serial))).isoformat()
        elif isinstance(excel_serial, dt.datetime):
            fecha_corte = excel_serial.date().isoformat()
        elif isinstance(excel_serial, dt.date):
            fecha_corte = excel_serial.isoformat()
        else:
            fecha_corte = dt.date.today().isoformat()

        agg = {}
        for row in all_rows[1:]:
            if not row or len(row) <= last_col:
                continue
            estado = row[4]
            cat = str(row[last_col] or '').strip().upper()
            if not estado:
                continue
            if estado not in agg:
                agg[estado] = {'total': 0, 'D0':0, 'D1':0, 'D2':0, 'D3':0, 'D4':0, 'NONE':0}
            agg[estado]['total'] += 1
            if cat in ('D0','D1','D2','D3','D4'):
                agg[estado][cat] += 1
            else:
                agg[estado]['NONE'] += 1

        rows_out = []
        for estado, c in agg.items():
            if c['total'] == 0: continue
            def pct(k): return round(100.0 * c[k] / c['total'], 2)
            dominante = 'Sin sequia'
            for nivel_nombre, k in [('Excepcional','D4'),('Extrema','D3'),('Severa','D2'),
                                     ('Moderada','D1'),('Anormal','D0')]:
                if pct(k) > 25:
                    dominante = nivel_nombre
                    break
            rows_out.append({
                'fecha_corte':           fecha_corte,
                'estado':                str(estado)[:80],
                # nivel_sequia es generated column en DB - NO enviarlo
                'pct_anomalo_seco':      pct('D0'),
                'pct_sequia_moderada':   pct('D1'),
                'pct_sequia_severa':     pct('D2'),
                'pct_sequia_extrema':    pct('D3'),
                'pct_sequia_excepcional': pct('D4'),
                'fuente':                'CONAGUA SMN Monitor de Sequia',
            })

        if not rows_out:
            print("   [WARN] No se agregaron filas.")
            return
        # FIX anti-zombie: UPSERT en (fecha_corte, estado). Si falla, NUNCA queda vacio.
        n = supa_post('monitor_sequia', rows_out, on_conflict='fecha_corte,estado')
        print(f"   [OK] {n} estados de sequia (corte {fecha_corte}, upsert)")
    except Exception as e:
        print(f"   [ERR] {e}")


# =====================================================================
# 4) PRESAS - SINA JSON con verify=False y timeout largo
# =====================================================================
def scrape_presas_cuencas():
    print("\n[PRESAS] SINA (JSON publico)...")
    url = "https://sinav30.conagua.gob.mx:8080/Common/PresasPrincipales/?valores=true"
    rows = []
    hoy = dt.date.today().isoformat()
    try:
        r = requests.get(url, headers=HEADERS, timeout=120, verify=False)
        if not r.ok:
            print(f"   [WARN] HTTP {r.status_code}")
            return
        try:
            data = r.json()
        except Exception:
            print(f"   [WARN] No es JSON: {r.text[:200]}")
            return
        arr = data if isinstance(data, list) else data.get('presas', [])
        print(f"   {len(arr)} presas recibidas del API")
        for p in arr:
            cap = _num(p.get('capacidadNAMO') or p.get('capacidad'))
            alm = _num(p.get('almacenaactual') or p.get('almacenamiento'))
            pct = _num(p.get('llenano') or p.get('porcentaje'))
            # FIX 2026-06-03: skipear bad data del API (0/0/0 = SINA no tiene reading hoy)
            # Diferente de presa REALMENTE vacia: si capacidad>0 pero alm=0 y pct=0, probable bad data.
            if (alm is None or alm == 0) and (pct is None or pct == 0):
                continue
            rows.append({
                'fecha_corte':            hoy,
                'presa':                  (p.get('nombreoficial') or p.get('nombre') or '')[:200],
                'estado':                 (p.get('estado') or '')[:80],
                'capacidad_total_hm3':    cap,
                'almacenamiento_hm3':     alm,
                'pct_almacenamiento':     pct,
                'region_hidrologica':     (p.get('region') or '')[:80],
                'fuente':                 'SINA CONAGUA',
            })
        if not rows:
            print("   [WARN] Sin presas extraidas.")
            return
        # FIX anti-zombie: UPSERT en (fecha_corte, presa). Si falla, NUNCA queda vacio.
        n = supa_post('presas_cuencas', rows, on_conflict='fecha_corte,presa')
        print(f"   [OK] {n} presas guardadas (upsert)")
    except Exception as e:
        print(f"   [ERR] {e}")


# =====================================================================
# MAIN
# =====================================================================
def main():
    print("Tequio CONAGUA Scraper v4 (fixes especificos + anti-zombie 2026-06-03)")
    print(f"   Supabase: {SUPABASE_URL}")
    scrape_clima_municipal()
    scrape_alertas_meteo()
    scrape_monitor_sequia()
    scrape_presas_cuencas()
    print("\nCONAGUA v4 scrape completo.")


if __name__ == '__main__':
    main()
