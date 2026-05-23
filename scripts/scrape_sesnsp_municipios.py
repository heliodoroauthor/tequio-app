#!/usr/bin/env python3
"""
Tequio - Scraper SESNSP Incidencia Delictiva Municipal.

SESNSP migro los datos abiertos a SharePoint en ZIP que contiene N archivos XLSX
(uno por anio: 2015.xlsx, 2016.xlsx, ..., 2025.xlsx).

Env vars:
  SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY
Opcional:
  GITHUB_RUN_ID
  SESNSP_ANIOS - comma-separated, filtra archivos del ZIP (mucho mas rapido)
  SESNSP_CSV_URL - URL directa (override del scraping)
"""
import csv
import io
import json
import os
import re
import sys
import unicodedata
import zipfile
from datetime import datetime, timezone
import requests
from bs4 import BeautifulSoup
try:
    import cloudscraper
    _CS = cloudscraper.create_scraper(browser={"browser":"chrome","platform":"windows","mobile":False})
    print("[sesnsp] cloudscraper activo para bypass Imperva")
except Exception as _e:
    import requests as _CS
    print(f"[sesnsp] cloudscraper NO disponible ({_e}); usando requests")

SB_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SB_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
GH_RUN_ID = os.environ.get("GITHUB_RUN_ID", "")
SESNSP_ANIOS = os.environ.get("SESNSP_ANIOS", "").strip()
SESNSP_CSV_URL_OVERRIDE = os.environ.get("SESNSP_CSV_URL", "").strip()

if not SB_URL or not SB_KEY:
    print("[sesnsp] Faltan SUPABASE_URL/SUPABASE_SERVICE_ROLE_KEY", file=sys.stderr)
    sys.exit(1)

SCRAPER_SLUG = "sesnsp_municipios"
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
SESNSP_PAGE = "https://www.gob.mx/sesnsp/acciones-y-programas/datos-abiertos-de-incidencia-delictiva"

ANIOS_FILTRO = set()
if SESNSP_ANIOS:
    ANIOS_FILTRO = {int(a.strip()) for a in SESNSP_ANIOS.split(",") if a.strip()}


def deaccent(s):
    if not s:
        return ""
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def normkey(s):
    s = deaccent(s or "").lower()
    return "".join(ch for ch in s if ch.isalnum())


def force_sharepoint_download(url):
    if "sharepoint.com" not in url:
        return url
    if "download=1" in url:
        return url
    sep = "&" if "?" in url else "?"
    return f"{url}{sep}download=1"


def find_dataset_url(html):
    soup = BeautifulSoup(html, "html.parser")
    anchors = soup.find_all("a", href=True)
    candidates = []
    for a in anchors:
        text = (a.get_text() or "").strip()
        t_norm = deaccent(text).lower()
        if ("fuero" in t_norm and "delitos" in t_norm and "municipal" in t_norm
                and "victima" not in t_norm and "victimas" not in t_norm
                and "tablero" not in t_norm):
            candidates.append((a["href"], text))
    print(f"[sesnsp] {len(candidates)} candidatos:")
    for h, t in candidates[:10]:
        print(f"  - {t[:80]!r}: {h[:100]}")
    if not candidates:
        # FIX-40 dbg: log all anchors to BD so we can see what changed
        try:
            sample = {"n_anchors": len(anchors), "html_head": html[:500], "html_tail": html[-500:], "anchors_first10": [(((a.get_text() or "").strip())[:80], (a.get("href","") or "")[:150]) for a in anchors[:10]]}
            _payload = [{
                "scraper_slug": "sesnsp_municipios",
                "status": "partial",
                "rows_inserted": 0,
                "rows_updated": 0,
                "rows_skipped": 0,
                "fuente_url": "https://www.gob.mx/sesnsp/acciones-y-programas/datos-abiertos-de-incidencia-delictiva",
                "http_status": 200,
                "error_msg": "DEBUG: anchors sample",
                "notes": json.dumps(sample, ensure_ascii=False)[:2500],
                "started_at": datetime.now(timezone.utc).isoformat(),
                "finished_at": datetime.now(timezone.utc).isoformat(),
            }]
            requests.post(f"{SB_URL}/rest/v1/scraper_logs", headers={"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}", "Content-Type":"application/json", "Prefer":"return=minimal"}, json=_payload, timeout=10)
        except Exception as _e:
            print(f"[sesnsp dbg log err] {_e}")
        return None, None

    def score(text):
        match = re.search(r"(\d{4})\s*-\s*(\d{4})", text)
        if match:
            return int(match.group(2)) - int(match.group(1))
        return 0

    candidates.sort(key=lambda x: score(x[1]), reverse=True)
    url, label = candidates[0]
    return force_sharepoint_download(url), label


def discover_dataset_url():
    if SESNSP_CSV_URL_OVERRIDE:
        url = force_sharepoint_download(SESNSP_CSV_URL_OVERRIDE)
        print(f"[sesnsp] usando override URL: {url}")
        return url, "OVERRIDE"
    print(f"[sesnsp] descubriendo URL desde {SESNSP_PAGE}")
    r = _CS.get(SESNSP_PAGE, headers={"User-Agent": UA}, timeout=60)
    r.raise_for_status()
    print(f"[sesnsp] HTML len={len(r.text)} status={r.status_code}")
    url, label = find_dataset_url(r.text)
    if not url:
        raise RuntimeError("No se encontro link SharePoint con 'Fuero comun - Delitos - municipal'")
    print(f"[sesnsp] dataset elegido: {label[:120]!r}")
    return url, label


def download_file(url):
    print(f"[sesnsp] descargando {url[:120]}")
    r = requests.get(url, headers={"User-Agent": UA, "Accept": "*/*"}, timeout=600, allow_redirects=True)
    r.raise_for_status()
    print(f"[sesnsp] bytes={len(r.content)} content-type={r.headers.get('Content-Type','')}")
    return r.content


def parse_xlsx_blob(blob):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    sheet = wb.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([("" if v is None else str(v)) for v in row])
    wb.close()
    return rows


def parse_csv_text(text):
    sample = text[:5000]
    sep = "," if sample.count(",") > sample.count(";") else ";"
    reader = csv.reader(io.StringIO(text), delimiter=sep)
    return list(reader)


def blob_to_text(blob):
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return blob.decode(enc)
        except UnicodeDecodeError:
            continue
    return blob.decode("latin-1")


MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "setiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}


def parse_int(s):
    if s is None:
        return None
    s = str(s).replace(",", "").strip()
    if s == "" or s.lower() in ("nan", "none", "null"):
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def unpivot(all_rows):
    if not all_rows:
        return []
    header_idx = 0
    for i, row in enumerate(all_rows[:5]):
        joined = " ".join(str(c) for c in row).lower()
        if "ano" in normkey(joined) or "anio" in normkey(joined) or "clave" in normkey(joined):
            header_idx = i
            break

    raw_headers = [str(c or "").strip() for c in all_rows[header_idx]]
    nh = [normkey(h) for h in raw_headers]

    def idx(*names):
        for n in names:
            if n in nh:
                return nh.index(n)
        return -1

    i_anio = idx("ano", "anio")
    i_cve_ent = idx("claveent", "cveent", "cveagee")
    i_entidad = idx("entidad", "estado")
    i_cve_mun = idx("cvemunicipio", "cveagemmunicipio", "cveagem", "clavemunicipio")
    i_municipio = idx("municipio", "nommun")
    i_bien = idx("bienjuridicoafectado", "bienjuridico")
    i_tipo = idx("tipodedelito", "tipodelito")
    i_subtipo = idx("subtipodedelito", "subtipodelito")
    i_modalidad = idx("modalidad")

    mes_indices = {}
    for i, h in enumerate(nh):
        mes_key = MESES.get(deaccent(h).lower())
        if mes_key:
            mes_indices[i] = mes_key

    if i_anio == -1 or i_cve_mun == -1:
        raise RuntimeError(f"Headers faltantes. Headers crudos: {raw_headers[:20]}")

    rows_long = []
    skipped = 0
    for row in all_rows[header_idx + 1:]:
        if not row or len([c for c in row if str(c or "").strip()]) < 5:
            continue
        try:
            anio = parse_int(row[i_anio])
            cve_ent_raw = str(row[i_cve_ent] or "").strip() if i_cve_ent >= 0 else ""
            cve_mun_raw = str(row[i_cve_mun] or "").strip() if i_cve_mun >= 0 else ""
            if not anio or not cve_mun_raw:
                skipped += 1
                continue
            cve_ent = cve_ent_raw.zfill(2)
            cve_mun = cve_mun_raw.zfill(3)
            if len(cve_mun) > 3:
                cve_mun = cve_mun[-3:]
            clave_inegi = f"{cve_ent}{cve_mun}"

            bien = str(row[i_bien] or "").strip() if i_bien >= 0 else None
            tipo = str(row[i_tipo] or "").strip() if i_tipo >= 0 else None
            subtipo = str(row[i_subtipo] or "").strip() if i_subtipo >= 0 else None
            modalidad = str(row[i_modalidad] or "").strip() if i_modalidad >= 0 else None
            entidad = str(row[i_entidad] or "").strip() if i_entidad >= 0 else None
            municipio = str(row[i_municipio] or "").strip() if i_municipio >= 0 else None

            for col_idx, mes_num in mes_indices.items():
                if col_idx >= len(row):
                    continue
                cantidad = parse_int(row[col_idx])
                if cantidad is None or cantidad == 0:
                    continue
                rows_long.append({
                    "clave_inegi": clave_inegi,
                    "clave_entidad": cve_ent,
                    "clave_municipio": cve_mun,
                    "nombre_estado": entidad or None,
                    "nombre_municipio": municipio or None,
                    "anio": anio,
                    "mes": mes_num,
                    "bien_juridico": bien or None,
                    "tipo_delito": tipo or None,
                    "subtipo_delito": subtipo or None,
                    "modalidad": modalidad or None,
                    "cantidad": cantidad,
                })
        except Exception:
            skipped += 1

    return rows_long, skipped


def sb_upsert(table, rows, on_conflict, batch_size=500):
    if not rows:
        return 0
    total = 0
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        url = f"{SB_URL}/rest/v1/{table}?on_conflict={on_conflict}"
        r = requests.post(url, headers={
            "Content-Type": "application/json", "apikey": SB_KEY,
            "Authorization": f"Bearer {SB_KEY}",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        }, data=json.dumps(batch), timeout=120)
        if not r.ok:
            raise RuntimeError(f"upsert batch {i//batch_size} status={r.status_code}: {r.text[:300]}")
        total += len(batch)
        if total % 10000 == 0:
            print(f"[sesnsp]   upserted {total}/{len(rows)}")
    return total


def log_scraper(status, summary, error_msg, started_at, fuente_url):
    try:
        requests.post(f"{SB_URL}/rest/v1/scraper_logs", headers={
            "Content-Type": "application/json", "apikey": SB_KEY,
            "Authorization": f"Bearer {SB_KEY}", "Prefer": "return=minimal",
        }, data=json.dumps([{
            "scraper_slug": SCRAPER_SLUG, "workflow_run_id": GH_RUN_ID or None,
            "status": status, "rows_inserted": summary.get("inserted", 0),
            "rows_updated": 0, "rows_skipped": summary.get("skipped", 0),
            "fuente_url": (fuente_url or SESNSP_PAGE)[:500],
            "http_status": 200 if status == "ok" else 500,
            "error_msg": (error_msg or "")[:1000] or None,
            "notes": json.dumps(summary)[:1000],
            "started_at": started_at,
            "finished_at": datetime.now(timezone.utc).isoformat(),
        }]), timeout=30)
    except Exception as exc:
        print(f"[sesnsp] no log: {exc}", file=sys.stderr)


def process_xlsx_blob(blob, anio_hint=None):
    """Devuelve lista de rows long para upsert."""
    all_rows = parse_xlsx_blob(blob)
    rows, skipped = unpivot(all_rows)
    return rows, skipped


def year_from_filename(name):
    m = re.search(r"(\d{4})", os.path.basename(name))
    return int(m.group(1)) if m else None


def main():
    started_at = datetime.now(timezone.utc).isoformat()
    summary = {"inserted": 0, "skipped": 0, "errors": [], "by_anio": {}}
    dataset_url = ""

    try:
        dataset_url, label = discover_dataset_url()
        blob = download_file(dataset_url)

        # SESNSP ZIP contiene 1 XLSX por anio
        if blob[:2] != b"PK":
            raise RuntimeError(f"Esperaba ZIP (PK header), recibi {blob[:8]!r}")

        with zipfile.ZipFile(io.BytesIO(blob)) as zf:
            names = zf.namelist()
            print(f"[sesnsp] ZIP contiene {len(names)} entries")
            xlsx_files = [n for n in names if n.lower().endswith(".xlsx")]
            print(f"[sesnsp] {len(xlsx_files)} archivos XLSX")

            for name in sorted(xlsx_files):
                file_year = year_from_filename(name)
                if ANIOS_FILTRO and file_year and file_year not in ANIOS_FILTRO:
                    print(f"[sesnsp] SKIP {name} (anio {file_year} no esta en filtro)")
                    continue

                print(f"[sesnsp] procesando {name} (anio detectado={file_year})")
                inner = zf.read(name)
                try:
                    rows, skipped = process_xlsx_blob(inner, anio_hint=file_year)
                except Exception as exc:
                    summary["errors"].append(f"{name}: {exc}")
                    print(f"[sesnsp]   FAIL parsing: {exc}", file=sys.stderr)
                    continue

                if ANIOS_FILTRO:
                    before = len(rows)
                    rows = [r for r in rows if r["anio"] in ANIOS_FILTRO]
                    if before != len(rows):
                        print(f"[sesnsp]   filtrado anio: {before} -> {len(rows)}")

                # Dedup
                seen = set()
                deduped = []
                for r in rows:
                    k = (r["clave_inegi"], r["anio"], r["mes"], r.get("bien_juridico"),
                         r.get("tipo_delito"), r.get("subtipo_delito"), r.get("modalidad"))
                    if k in seen:
                        summary["skipped"] += 1
                        continue
                    seen.add(k)
                    deduped.append(r)

                if not deduped:
                    print(f"[sesnsp]   sin filas para upsert")
                    continue

                print(f"[sesnsp]   upserting {len(deduped)} filas...")
                ins = sb_upsert("delitos_municipios", deduped,
                                on_conflict="clave_inegi,anio,mes,bien_juridico,tipo_delito,subtipo_delito,modalidad")
                summary["inserted"] += ins
                summary["by_anio"][str(file_year)] = ins
                print(f"[sesnsp]   {name}: +{ins} filas (total: {summary['inserted']})")

        status = "ok" if summary["inserted"] > 0 else ("partial" if summary["errors"] else "fail")
    except Exception as exc:
        import traceback; traceback.print_exc()
        summary["errors"].append(str(exc))
        status = "fail"

    log_scraper(status, summary, "; ".join(summary["errors"][:5]) or None, started_at, dataset_url)
    print(f"[sesnsp] DONE status={status} total={summary['inserted']} by_anio={summary['by_anio']}")
    sys.exit(0 if status in ("ok", "partial") and summary["inserted"] > 0 else 1)


if __name__ == "__main__":
    main()
