#!/usr/bin/env python3
"""Tequio - Scraper INEGI EFIPEM Finanzas Municipales (anual)."""
import csv
import io
import json
import os
import re
import sys
import unicodedata
import zipfile
from datetime import datetime, timezone
import requests

SB_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SB_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
GH_RUN_ID = os.environ.get("GITHUB_RUN_ID", "")
URL_OVERRIDE = os.environ.get("EFIPEM_URL", "").strip()
ANIOS_FILTRO = os.environ.get("EFIPEM_ANIOS", "").strip()

if not SB_URL or not SB_KEY:
    print("[efipem] Faltan SUPABASE_URL/SUPABASE_SERVICE_ROLE_KEY", file=sys.stderr)
    sys.exit(1)

SCRAPER_SLUG = "inegi_efipem"
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
EFIPEM_ZIP = "https://www.inegi.org.mx/contenidos/programas/finanzas/datosabiertos/efipem.zip"

ANIOS_SET = set()
if ANIOS_FILTRO:
    ANIOS_SET = {int(a.strip()) for a in ANIOS_FILTRO.split(",") if a.strip()}


def deaccent(s):
    if not s: return ""
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def normkey(s):
    s = deaccent(s or "").lower()
    return "".join(ch for ch in s if ch.isalnum())


def parse_num(v):
    if v is None or v == "": return None
    s = str(v).replace(",", "").replace("$", "").strip()
    if not s or s.lower() in ("nan", "null", "none", "n/d", "n.d.", "-"):
        return None
    try: return float(s)
    except: return None


def parse_int(v):
    n = parse_num(v)
    if n is None: return None
    try: return int(n)
    except: return None


def pick(rec, *keys):
    for k in keys:
        v = rec.get(k)
        if v not in (None, ""): return v
    return None


NOMBRES_ESTADO = {
    "01": "Aguascalientes", "02": "Baja California", "03": "Baja California Sur",
    "04": "Campeche", "05": "Coahuila", "06": "Colima", "07": "Chiapas",
    "08": "Chihuahua", "09": "Ciudad de Mexico", "10": "Durango",
    "11": "Guanajuato", "12": "Guerrero", "13": "Hidalgo", "14": "Jalisco",
    "15": "Estado de Mexico", "16": "Michoacan", "17": "Morelos", "18": "Nayarit",
    "19": "Nuevo Leon", "20": "Oaxaca", "21": "Puebla", "22": "Queretaro",
    "23": "Quintana Roo", "24": "San Luis Potosi", "25": "Sinaloa", "26": "Sonora",
    "27": "Tabasco", "28": "Tamaulipas", "29": "Tlaxcala", "30": "Veracruz",
    "31": "Yucatan", "32": "Zacatecas",
}


def download_zip():
    url = URL_OVERRIDE or EFIPEM_ZIP
    print(f"[efipem] descargando {url}")
    r = requests.get(url, headers={"User-Agent": UA, "Accept": "*/*"}, timeout=600)
    r.raise_for_status()
    print(f"[efipem] bytes={len(r.content)} content-type={r.headers.get('Content-Type','')}")
    return r.content, url


def normalize_capitulo(rubro):
    if not rubro: return None
    t = deaccent(rubro).lower()
    if "impuesto" in t: return "impuestos"
    if "derecho" in t: return "derechos"
    if "producto" in t: return "productos"
    if "aprovecham" in t: return "aprovechamientos"
    if "participacion" in t: return "participaciones_federales"
    if "aportacion" in t: return "aportaciones_federales"
    if "convenio" in t: return "convenios"
    if "financiamiento" in t or "deuda" in t: return "financiamiento"
    if "transferencia" in t: return "transferencias"
    if "servicio personal" in t or "personal" in t: return "servicios_personales"
    if "material" in t or "suministro" in t: return "materiales"
    if "servicios general" in t: return "servicios_generales"
    if "obra publica" in t or "inversion" in t: return "obra_publica_inversion"
    if "subsidio" in t or "ayuda" in t: return "subsidios_ayudas"
    if "bienes muebles" in t or "inmueble" in t: return "bienes_muebles_inmuebles"
    return None


def detect_flujo(concepto):
    t = deaccent(concepto or "").lower()
    if "ingreso" in t: return "ingreso"
    if "egreso" in t or "gasto" in t: return "egreso"
    return None


def map_row(rec, year_hint=None):
    cve_ent_raw = pick(rec, "id_entidad", "cve_entidad", "cve_ent", "ent", "claveent", "cveentidad", "identidad")
    cve_mun_raw = pick(rec, "id_municipio", "cve_municipio", "cve_mun", "mun", "clavemunicipio", "cvemunicipio", "idmunicipio")
    if not cve_ent_raw or not cve_mun_raw:
        return None
    cve_ent = str(cve_ent_raw).strip().zfill(2)
    if len(cve_ent) > 2: cve_ent = cve_ent[-2:]
    cve_mun = str(cve_mun_raw).strip().zfill(3)
    if len(cve_mun) > 3: cve_mun = cve_mun[-3:]
    clave_inegi = f"{cve_ent}{cve_mun}"
    if cve_mun == "000":
        return None  # estado nivel

    nombre_estado = NOMBRES_ESTADO.get(cve_ent) or (pick(rec, "entidad", "nom_entidad") or "").strip() or None
    nombre_municipio = (pick(rec, "municipio", "nom_municipio", "nommun") or "").strip().title() or None

    anio = parse_int(pick(rec, "anio", "ano", "ciclo", "year"))
    if not anio and year_hint:
        anio = year_hint
    if not anio:
        return None

    concepto = (pick(rec, "rubro", "concepto", "descripcion", "capitulo", "descapitulo", "estimacion") or "").strip() or None
    flujo = detect_flujo(concepto)
    # Heuristic: si concepto incluye "ingresos brutos" o "egresos brutos"
    if not flujo:
        flujo = "ingreso"

    capitulo = normalize_capitulo(concepto)

    monto = parse_num(pick(rec, "valor", "monto", "importe", "total"))
    if monto is None:
        return None

    return {
        "clave_inegi": clave_inegi,
        "clave_entidad": cve_ent,
        "clave_municipio": cve_mun,
        "nombre_estado": nombre_estado,
        "nombre_municipio": nombre_municipio,
        "anio": anio,
        "flujo": flujo,
        "capitulo": capitulo,
        "concepto": concepto[:500] if concepto else None,
        "monto": monto,
        "unidad": "MXN",
        "fuente": "INEGI EFIPEM",
        "fuente_url": EFIPEM_ZIP,
        "fecha_extraccion": datetime.now(timezone.utc).date().isoformat(),
    }


def sb_upsert(table, rows, on_conflict, batch_size=500):
    if not rows: return 0
    total = 0
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        url = f"{SB_URL}/rest/v1/{table}?on_conflict={on_conflict}"
        r = requests.post(url, headers={
            "Content-Type": "application/json", "apikey": SB_KEY,
            "Authorization": f"Bearer {SB_KEY}",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        }, data=json.dumps(batch), timeout=120)
        if not r.ok:
            raise RuntimeError(f"upsert batch {i//batch_size} status={r.status_code}: {r.text[:300]}")
        total += len(batch)
        if total % 10000 == 0 or total == len(rows):
            print(f"[efipem] upserted {total}/{len(rows)}")
    return total


def log_scraper(status, summary, error_msg, started_at, fuente_url):
    try:
        requests.post(f"{SB_URL}/rest/v1/scraper_logs", headers={
            "Content-Type": "application/json", "apikey": SB_KEY,
            "Authorization": f"Bearer {SB_KEY}", "Prefer": "return=minimal",
        }, data=json.dumps([{
            "scraper_slug": SCRAPER_SLUG, "workflow_run_id": GH_RUN_ID or None,
            "status": status, "rows_inserted": summary.get("inserted", 0),
            "rows_updated": 0, "rows_skipped": summary.get("skipped", 0),
            "fuente_url": (fuente_url or EFIPEM_ZIP)[:500],
            "http_status": 200 if status == "ok" else 500,
            "error_msg": (error_msg or "")[:1000] or None,
            "notes": json.dumps(summary)[:1000],
            "started_at": started_at,
            "finished_at": datetime.now(timezone.utc).isoformat(),
        }]), timeout=30)
    except Exception as exc:
        print(f"[efipem] no log: {exc}", file=sys.stderr)


def parse_csv_text(text):
    sample = text[:5000]
    sep = "," if sample.count(",") > sample.count(";") else ";"
    reader = csv.reader(io.StringIO(text), delimiter=sep)
    return list(reader)


def parse_xlsx_blob(blob):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    sheet = wb.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([("" if v is None else str(v)) for v in row])
    wb.close()
    return rows


def blob_to_text(blob):
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return blob.decode(enc)
        except UnicodeDecodeError:
            continue
    return blob.decode("latin-1")


def year_from_filename(name):
    m = re.search(r"(20\d{2})", name)
    return int(m.group(1)) if m else None


def process_blob(blob, ext, year_hint=None):
    """Parsea CSV/TXT/XLSX y devuelve filas mapeadas."""
    ext = ext.lower()
    if ext in (".xlsx", ".xls"):
        rows_raw = parse_xlsx_blob(blob)
    else:
        # CSV or TXT — INEGI a veces usa TXT con tab/coma
        text = blob_to_text(blob)
        rows_raw = parse_csv_text(text)

    if not rows_raw: return []
    raw_headers = [str(c or "").strip() for c in rows_raw[0]]
    nh = [normkey(h) for h in raw_headers]
    print(f"[efipem]   headers ({len(raw_headers)}): {raw_headers[:15]}")

    out = []
    for row in rows_raw[1:]:
        if not row or not any(str(c or "").strip() for c in row): continue
        rec = dict(zip(nh, [str(c or "").strip() for c in row]))
        m = map_row(rec, year_hint=year_hint)
        if m:
            if ANIOS_SET and m["anio"] not in ANIOS_SET:
                continue
            out.append(m)
    return out


def main():
    started_at = datetime.now(timezone.utc).isoformat()
    summary = {"inserted": 0, "skipped": 0, "errors": [], "files": {}, "zip_contents": []}
    src_url = ""
    try:
        blob, src_url = download_zip()
        if blob[:2] != b"PK":
            raise RuntimeError(f"No es ZIP (header={blob[:8]!r})")

        with zipfile.ZipFile(io.BytesIO(blob)) as zf:
            names = zf.namelist()
            print(f"[efipem] ZIP contiene {len(names)} archivos:")
            for n in names:
                print(f"  - {n}")
            summary["zip_contents"] = names

            # Procesar todos los archivos con extensiones procesables
            all_mapped = []
            for name in names:
                if name.endswith("/"): continue
                ext = os.path.splitext(name)[1].lower() or ""
                if ext not in (".csv", ".txt", ".xlsx", ".xls"):
                    print(f"[efipem] SKIP {name} (ext={ext})")
                    continue
                year_hint = year_from_filename(name)
                if ANIOS_SET and year_hint and year_hint not in ANIOS_SET:
                    print(f"[efipem] SKIP {name} (anio {year_hint} fuera de filtro)")
                    continue
                print(f"[efipem] procesando {name} (year_hint={year_hint}, ext={ext})")
                try:
                    inner = zf.read(name)
                    rows = process_blob(inner, ext, year_hint=year_hint)
                    summary["files"][name] = len(rows)
                    all_mapped.extend(rows)
                    print(f"[efipem]   {name}: +{len(rows)} rows")
                except Exception as exc:
                    summary["errors"].append(f"{name}: {exc}")
                    print(f"[efipem]   FAIL {name}: {exc}", file=sys.stderr)

        print(f"[efipem] total mapeados: {len(all_mapped)}")
        if all_mapped: print(f"[efipem] sample: {all_mapped[0]}")

        seen = set()
        deduped = []
        for m in all_mapped:
            k = (m["clave_inegi"], m["anio"], m["flujo"], m.get("capitulo"), m.get("concepto"))
            if k in seen:
                summary["skipped"] += 1
                continue
            seen.add(k)
            deduped.append(m)
        print(f"[efipem] dedup: {len(all_mapped)} -> {len(deduped)}")

        if deduped:
            ins = sb_upsert("finanzas_municipales_inegi", deduped,
                            on_conflict="clave_inegi,anio,flujo,capitulo,concepto")
            summary["inserted"] = ins
            status = "ok"
        else:
            status = "fail"
    except Exception as exc:
        import traceback; traceback.print_exc()
        summary["errors"].append(str(exc))
        status = "fail"

    log_scraper(status, summary, "; ".join(summary["errors"][:5]) or None, started_at, src_url)
    print(f"[efipem] DONE status={status} total={summary['inserted']}")
    sys.exit(0 if status == "ok" else 1)


if __name__ == "__main__":
    main()
