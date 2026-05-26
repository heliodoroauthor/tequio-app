#!/usr/bin/env python3
"""
scrape_sequia.py — Monitor de Sequía CONAGUA SMN (extracto de v4)

Sólo scrapea monitor_sequia (no clima, alertas ni presas).
CONAGUA SMN publica el XLSX cada 2 semanas (día 15 y último de cada mes).

XLSX wide-format:
  Cols 0-8: metadatos (CVE_ENT, CVE_MUN, NOMBRE_MUN, ENTIDAD, etc)
  Cols 9+:  cada columna = una fecha (Excel serial)
  Filas:    2,479 municipios
  Valores:  '' (sin sequía) | D0 | D1 | D2 | D3 | D4

Setup:
    pip install requests openpyxl cloudscraper urllib3

Uso:
    export SUPABASE_URL=...
    export SUPABASE_SERVICE_ROLE_KEY=...
    python scrape_sequia.py

🦎 Cero Invención · Tequio · 2026
"""

import os
import sys
import io
import datetime as dt
import requests
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Bypass del cert chain roto de SMN
import ssl as _ssl
from requests.adapters import HTTPAdapter as _HA
try:
    from urllib3.util.ssl_ import create_urllib3_context as _ctx
except ImportError:
    _ctx = None


class _NoVerify(_HA):
    def init_poolmanager(self, *a, **kw):
        if _ctx:
            c = _ctx()
            c.check_hostname = False
            c.verify_mode = _ssl.CERT_NONE
            kw['ssl_context'] = c
        return super().init_poolmanager(*a, **kw)


try:
    import cloudscraper
    _CS = cloudscraper.create_scraper(
        browser={'browser': 'chrome', 'platform': 'darwin', 'mobile': False}
    )
    _CS.verify = False
    _CS.mount('https://', _NoVerify())

    def gob_get(url, **kw):
        kw.setdefault('verify', False)
        return _CS.get(url, **kw)
    print("  [cloudscraper] OK — anti-bot gob.mx + NoVerifyAdapter activo")
except ImportError:
    _S = requests.Session()
    _S.mount('https://', _NoVerify())
    _S.verify = False

    def gob_get(url, **kw):
        kw.setdefault('verify', False)
        return _S.get(url, **kw)
    print("  [cloudscraper] NO disponible — usando requests con NoVerifyAdapter")


SUPABASE_URL = os.environ['SUPABASE_URL'].rstrip('/')
SERVICE_KEY = os.environ['SUPABASE_SERVICE_ROLE_KEY']

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (compatible; TequioCivicBot/4.0; +https://tequio.app)',
    'Accept': 'application/octet-stream,*/*',
    'Accept-Language': 'es-MX,es;q=0.9',
}
TIMEOUT = 90


def supa_upsert(table, rows, on_conflict):
    if not rows:
        return 0
    url = f"{SUPABASE_URL}/rest/v1/{table}?on_conflict={on_conflict}"
    r = requests.post(
        url, json=rows,
        headers={
            'Content-Type': 'application/json',
            'apikey': SERVICE_KEY,
            'Authorization': f'Bearer {SERVICE_KEY}',
            'Prefer': 'resolution=merge-duplicates,return=minimal',
        }, timeout=TIMEOUT,
    )
    if not r.ok:
        print(f"   [ERR] Supabase {table} {r.status_code}: {r.text[:300]}")
        return 0
    return len(rows)


def scrape_monitor_sequia():
    print("\n[SEQUIA] Monitor de Sequia (XLSX wide-format)...")
    url_xlsx = "https://smn.conagua.gob.mx/tools/RESOURCES/Monitor%20de%20Sequia%20en%20Mexico/MunicipiosSequia.xlsx"

    try:
        r = gob_get(url_xlsx, headers=HEADERS, timeout=TIMEOUT)
        if not r.ok:
            print(f"   [ERR] HTTP {r.status_code}")
            sys.exit(1)
        print(f"   Bajaron {len(r.content)} bytes del XLSX")
    except Exception as e:
        print(f"   [ERR] Fetch: {e}")
        sys.exit(1)

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("   [ERR] pip install openpyxl")
        sys.exit(1)

    wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    ws = wb['MUNICIPIOS'] if 'MUNICIPIOS' in wb.sheetnames else wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 10:
        print(f"   [ERR] Excel casi vacío: {len(all_rows)} filas")
        sys.exit(1)
    header = all_rows[0]
    FIRST_DATE_COL = 9

    # Buscar última columna con >100 valores no-vacíos
    last_col = FIRST_DATE_COL
    for col_idx in range(len(header) - 1, FIRST_DATE_COL - 1, -1):
        non_empty = sum(
            1 for row in all_rows[1:]
            if col_idx < len(row) and row[col_idx]
        )
        if non_empty > 100:
            last_col = col_idx
            break

    # Convertir Excel serial a ISO date
    excel_serial = header[last_col]
    if isinstance(excel_serial, (int, float)):
        fecha_corte = (
            dt.date(1899, 12, 30) + dt.timedelta(days=int(excel_serial))
        ).isoformat()
    elif isinstance(excel_serial, dt.datetime):
        fecha_corte = excel_serial.date().isoformat()
    elif isinstance(excel_serial, dt.date):
        fecha_corte = excel_serial.isoformat()
    else:
        fecha_corte = dt.date.today().isoformat()
    print(f"   Fecha de corte detectada: {fecha_corte}")

    # Agregar por estado
    agg = {}
    municipios_procesados = 0
    for row in all_rows[1:]:
        if not row or len(row) <= last_col:
            continue
        estado = row[4]  # ENTIDAD
        cat = str(row[last_col] or '').strip().upper()
        if not estado:
            continue
        if estado not in agg:
            agg[estado] = {'total': 0, 'D0': 0, 'D1': 0,
                           'D2': 0, 'D3': 0, 'D4': 0}
        agg[estado]['total'] += 1
        municipios_procesados += 1
        if cat in ('D0', 'D1', 'D2', 'D3', 'D4'):
            agg[estado][cat] += 1

    print(f"   {municipios_procesados} municipios procesados en {len(agg)} estados")

    rows_out = []
    for estado, c in agg.items():
        if c['total'] == 0:
            continue

        def pct(k):
            return round(100.0 * c[k] / c['total'], 2)

        rows_out.append({
            'fecha_corte':            fecha_corte,
            'estado':                 str(estado)[:80],
            # nivel_sequia es generated column en DB — no enviarlo
            'pct_anomalo_seco':       pct('D0'),
            'pct_sequia_moderada':    pct('D1'),
            'pct_sequia_severa':      pct('D2'),
            'pct_sequia_extrema':     pct('D3'),
            'pct_sequia_excepcional': pct('D4'),
            'fuente':                 'CONAGUA SMN Monitor de Sequia',
        })

    if not rows_out:
        print("   [ERR] No se agregaron filas")
        sys.exit(1)

    n = supa_upsert('monitor_sequia', rows_out, on_conflict='estado,fecha_corte')
    print(f"   [OK] {n} estados upserted (corte {fecha_corte})")
    print(f"\n=== RESUMEN ===\n  fecha:     {fecha_corte}\n  estados:   {n}\n  municipios: {municipios_procesados}")


if __name__ == '__main__':
    scrape_monitor_sequia()
