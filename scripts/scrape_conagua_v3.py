#!/usr/bin/env python3
"""
scrape_conagua_v3.py — HTML/CSV puros (sin gzip, sin endpoints oscuros)
=========================================================================

CAMBIOS RESPECTO v2:
  - clima_municipal:  Scrape HTML del pronostico general por estado.
  - monitor_sequia:   Descarga el Excel quincenal y agrega por ESTADO.
  - alertas_meteo:    Scrape HTML directo de la pagina de Avisos del SMN.
  - presas_cuencas:   JSON publico del SINA (sinav30.conagua.gob.mx).

REGLAS:
  * Si la pagina carga en navegador, el scraper la puede leer.
  * Preferimos CSV/texto plano sobre HTML cuando este disponible.
  * Si todo falla en un modulo, registramos error y seguimos.
  * Schemas alineados con Supabase actual (verificado mayo 2026).

Variables de entorno requeridas:
  SUPABASE_URL
  SUPABASE_SERVICE_ROLE_KEY
"""

import os
import re
import io
import datetime as dt
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin

SUPABASE_URL = os.environ['SUPABASE_URL'].rstrip('/')
SERVICE_KEY  = os.environ['SUPABASE_SERVICE_ROLE_KEY']

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (compatible; TequioCivicBot/3.0; +https://tequio.app)',
    'Accept': 'text/html,application/xhtml+xml,application/xml,text/plain,*/*',
    'Accept-Language': 'es-MX,es;q=0.9,en;q=0.8',
}
TIMEOUT = 60


# =====================================================================
# HELPERS
# =====================================================================
def supa_post(table, rows, on_conflict=None):
    if not rows:
        return 0
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if on_conflict:
        url += f"?on_conflict={on_conflict}"
        prefer = 'resolution=merge-duplicates,return=minimal'
    else:
        prefer = 'return=minimal'
    try:
        r = requests.post(url, json=rows, headers={
            'Content-Type': 'application/json',
            'apikey': SERVICE_KEY,
            'Authorization': f'Bearer {SERVICE_KEY}',
            'Prefer': prefer,
        }, timeout=TIMEOUT)
        if not r.ok:
            print(f"   [WARN] Supabase {table} {r.status_code}: {r.text[:250]}")
            return 0
        return len(rows)
    except Exception as e:
        print(f"   [WARN] Supabase {table} excepcion: {e}")
        return 0


def supa_truncate(table):
    """Borra todos los registros de una tabla (para refresh completo)."""
    try:
        r = requests.delete(
            f"{SUPABASE_URL}/rest/v1/{table}?id=gt.0",
            headers={
                'apikey': SERVICE_KEY,
                'Authorization': f'Bearer {SERVICE_KEY}',
                'Prefer': 'return=minimal',
            }, timeout=TIMEOUT)
        return r.ok
    except Exception:
        return False


def _num(v):
    if v is None: return None
    s = str(v).replace(',', '').strip()
    if s in ('', '-', 'N/D', 'NA', 'ND', 'S/D'): return None
    try: return float(s)
    except: return None


# =====================================================================
# 1) CLIMA MUNICIPAL — scrape HTML del pronostico general
# =====================================================================
def scrape_clima_municipal():
    """
    URL: https://smn.conagua.gob.mx/es/pronostico-meteorologico-general
    Schema:
      fecha_pronostico (date), municipio_id (text), municipio (text), estado (text),
      temp_max (num), temp_min (num), prob_lluvia (int), desc_cielo (text),
      velocidad_viento (num), direccion_viento (text), fuente (text)
    """
    print("\n[CLIMA] Pronostico general por estado (HTML)...")
    url = "https://smn.conagua.gob.mx/es/pronostico-meteorologico-general"
    try:
        r = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
        if not r.ok:
            print(f"   [WARN] HTTP {r.status_code} en {url}")
            return
        soup = BeautifulSoup(r.text, 'lxml')
        rows = []
        hoy = dt.date.today().isoformat()

        # Estrategia 1: Buscar todas las tablas y filtrar las que tienen 'estado'+'temp'
        for table in soup.find_all('table'):
            header_text = ' '.join(
                th.get_text(' ', strip=True).lower()
                for th in table.find_all(['th', 'td'])[:8]
            )
            if not any(k in header_text for k in ['estado', 'entidad', 'temp', 'max', 'min']):
                continue
            for tr in table.find_all('tr')[1:]:
                cells = [td.get_text(' ', strip=True) for td in tr.find_all('td')]
                if len(cells) < 3:
                    continue
                estado = cells[0].strip()
                if not estado or len(estado) > 80 or len(estado) < 3:
                    continue
                tmax = _num(cells[1]) if len(cells) > 1 else None
                tmin = _num(cells[2]) if len(cells) > 2 else None
                if tmax is None and tmin is None:
                    continue
                cielo = cells[3][:120] if len(cells) > 3 else ''
                rows.append({
                    'fecha_pronostico': hoy,
                    'municipio_id': '',  # nivel estatal
                    'municipio':    '',
                    'estado':       estado[:80],
                    'temp_max':     tmax,
                    'temp_min':     tmin,
                    'prob_lluvia':  None,
                    'desc_cielo':   cielo,
                    'velocidad_viento': None,
                    'direccion_viento': '',
                    'fuente':       'SMN CONAGUA',
                })

        if not rows:
            print("   [WARN] No se encontro tabla de pronostico en el HTML.")
            print("   [INFO] La pagina puede ser totalmente JS; necesitariamos Playwright.")
            return

        # Schema actual no tiene UNIQUE constraint, así que limpiamos primero
        supa_truncate('clima_municipal')
        n = supa_post('clima_municipal', rows)
        print(f"   [OK] {n} pronosticos estatales guardados")
    except Exception as e:
        print(f"   [ERR] {e}")


# =====================================================================
# 2) MONITOR DE SEQUIA — Excel quincenal vinculado en el HTML
# =====================================================================
def scrape_monitor_sequia():
    """
    URL: https://smn.conagua.gob.mx/es/climatologia/monitor-de-sequia/monitor-de-sequia-en-mexico
    Schema:
      fecha_corte (date), estado (text), nivel_sequia (text),
      pct_anomalo_seco, pct_sequia_moderada, pct_sequia_severa,
      pct_sequia_extrema, pct_sequia_excepcional (numeric),
      fuente (text)

    Estrategia: encontrar enlace al Excel mas reciente, descargarlo,
    agregar por estado.
    """
    print("\n[SEQUIA] Monitor de Sequia (Excel quincenal)...")
    url = "https://smn.conagua.gob.mx/es/climatologia/monitor-de-sequia/monitor-de-sequia-en-mexico"
    try:
        r = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
        if not r.ok:
            print(f"   [WARN] HTTP {r.status_code}")
            return
        soup = BeautifulSoup(r.text, 'lxml')

        # Buscar el primer .xlsx que tenga "MS" o "Monitor" en el texto/URL
        candidatos = []
        for a in soup.find_all('a', href=True):
            href = a['href']
            low_h = href.lower()
            low_t = a.get_text(' ', strip=True).lower()
            if low_h.endswith('.xlsx') or low_h.endswith('.xls'):
                score = 0
                if 'ms' in low_h or 'monitor' in low_h or 'sequia' in low_h: score += 2
                if 'monitor' in low_t or 'sequia' in low_t:                 score += 1
                full = urljoin(url, href)
                candidatos.append((score, full, a.get_text(strip=True)))

        if not candidatos:
            print("   [WARN] No se encontraron archivos Excel en la pagina.")
            return

        # Más score primero, luego asume el primer link es el mas reciente
        candidatos.sort(key=lambda x: (-x[0],))
        _, archivo_url, archivo_nombre = candidatos[0]
        print(f"   Descargando: {archivo_nombre}")
        r2 = requests.get(archivo_url, headers=HEADERS, timeout=TIMEOUT)
        if not r2.ok:
            print(f"   [WARN] HTTP {r2.status_code} al descargar")
            return

        try:
            from openpyxl import load_workbook
        except ImportError:
            print("   [WARN] openpyxl no instalado. Agregar a requirements.")
            return

        wb = load_workbook(io.BytesIO(r2.content), read_only=True, data_only=True)
        ws = wb.active

        # Detectar columnas. Esperamos: Entidad, D0, D1, D2, D3, D4 (anormal,moderada,severa,extrema,excepcional)
        # o similar. Probamos varios layouts.
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 2:
            print("   [WARN] Excel vacio.")
            return

        # Buscar fila de header (la que contiene 'estado' o 'entidad')
        header_idx = None
        for i, row in enumerate(all_rows[:15]):
            txt = ' '.join(str(c or '').lower() for c in row)
            if ('estado' in txt or 'entidad' in txt) and ('d0' in txt or 'sequia' in txt or 'd1' in txt):
                header_idx = i
                break

        if header_idx is None:
            print("   [WARN] No se encontro header en el Excel.")
            return

        header = [str(c or '').strip() for c in all_rows[header_idx]]
        rows_out = []
        fecha_corte = dt.date.today().isoformat()

        # Intentar extraer fecha del nombre del archivo (formato YYYYMMDD o MSYYYYMMDD)
        m_fecha = re.search(r'(\d{4})[-_]?(\d{2})[-_]?(\d{2})', archivo_url)
        if m_fecha:
            try:
                fecha_corte = f"{m_fecha.group(1)}-{m_fecha.group(2)}-{m_fecha.group(3)}"
            except:
                pass

        for row in all_rows[header_idx+1:]:
            rd = dict(zip(header, row))
            # Tomar estado de la primera columna no vacia
            estado = ''
            for c in row[:3]:
                if c and isinstance(c, str) and len(c.strip()) > 2:
                    estado = c.strip()
                    break
            if not estado or estado.lower() in ('total', 'mexico'):
                continue

            # Mapear columnas D0-D4 (% del estado en cada nivel)
            def find_col(*candidates):
                for k, v in rd.items():
                    kl = str(k or '').lower()
                    for cand in candidates:
                        if cand in kl:
                            return _num(v)
                return None

            row_data = {
                'fecha_corte':           fecha_corte,
                'estado':                estado[:80],
                'nivel_sequia':          '',  # se computa abajo
                'pct_anomalo_seco':      find_col('d0', 'anormal'),
                'pct_sequia_moderada':   find_col('d1', 'moderada'),
                'pct_sequia_severa':     find_col('d2', 'severa'),
                'pct_sequia_extrema':    find_col('d3', 'extrema'),
                'pct_sequia_excepcional': find_col('d4', 'excepcional'),
                'fuente':                'CONAGUA SMN ' + archivo_url[:100],
            }

            # Determinar nivel dominante
            niveles = [
                ('Excepcional', row_data['pct_sequia_excepcional']),
                ('Extrema',     row_data['pct_sequia_extrema']),
                ('Severa',      row_data['pct_sequia_severa']),
                ('Moderada',    row_data['pct_sequia_moderada']),
                ('Anormal',     row_data['pct_anomalo_seco']),
            ]
            for nombre, pct in niveles:
                if pct and pct > 25:
                    row_data['nivel_sequia'] = nombre
                    break
            if not row_data['nivel_sequia']:
                row_data['nivel_sequia'] = 'Sin sequia'

            rows_out.append(row_data)

        if not rows_out:
            print("   [WARN] Excel parseado pero sin filas validas.")
            return

        supa_truncate('monitor_sequia')
        n = supa_post('monitor_sequia', rows_out)
        print(f"   [OK] {n} estados de Monitor de Sequia guardados (corte {fecha_corte})")
    except Exception as e:
        print(f"   [ERR] {e}")


# =====================================================================
# 3) ALERTAS METEO — scrape HTML de avisos
# =====================================================================
def scrape_alertas_meteo():
    """
    URL: https://smn.conagua.gob.mx/es/avisos-tiempo
    Schema:
      id, tipo, nombre, nivel, zona_afectada, descripcion,
      vigente_desde (ts), vigente_hasta (ts), fuente, url_oficial, insertado_at
    """
    print("\n[ALERTAS] Avisos meteorologicos vigentes...")
    url = "https://smn.conagua.gob.mx/es/avisos-tiempo"
    try:
        r = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
        if not r.ok:
            print(f"   [WARN] HTTP {r.status_code}")
            return
        soup = BeautifulSoup(r.text, 'lxml')

        rows = []
        ahora_iso = dt.datetime.utcnow().isoformat()
        keywords_aviso = ('aviso', 'alerta', 'pronostico', 'huracan', 'tormenta',
                          'frente frio', 'lluvia', 'vientos', 'nortes', 'onda calor',
                          'ciclon', 'norte', 'depresion')

        for art in soup.select('article, .card, .aviso, .item, tr, .panel-body'):
            titulo_el = art.find(['h1', 'h2', 'h3', 'h4', 'strong'])
            titulo = titulo_el.get_text(strip=True) if titulo_el else ''
            cuerpo = art.get_text(' ', strip=True)
            if not cuerpo or len(cuerpo) < 60 or len(cuerpo) > 5000:
                continue
            low = cuerpo.lower()
            if not any(k in low for k in keywords_aviso):
                continue

            # Detectar nivel/severidad
            nivel = 'verde'
            if any(k in low for k in ['extremo', 'huracan', 'cat. 4', 'cat. 5']):
                nivel = 'rojo'
            elif any(k in low for k in ['severo', 'intenso', 'cat. 3']):
                nivel = 'naranja'
            elif any(k in low for k in ['fuerte', 'moderado', 'cat. 1', 'cat. 2']):
                nivel = 'amarillo'

            # Extraer zona si esta en el texto
            zona_match = re.search(r'(?:estado|estados|zona|region)[^.]{0,150}', low)
            zona = zona_match.group(0)[:200] if zona_match else ''

            rows.append({
                'tipo':           'aviso_smn',
                'nombre':         (titulo or cuerpo[:80])[:200],
                'nivel':          nivel,
                'zona_afectada':  zona,
                'descripcion':    cuerpo[:2000],
                'vigente_desde':  ahora_iso,
                'vigente_hasta':  None,
                'fuente':         'CONAGUA SMN',
                'url_oficial':    url,
            })

        # Dedupe por nombre+primeros 100 chars de descripcion
        seen = set()
        unique = []
        for row in rows:
            k = (row['nombre'][:80], row['descripcion'][:120])
            if k not in seen:
                seen.add(k)
                unique.append(row)

        if not unique:
            print("   [INFO] No hay avisos vigentes en este momento.")
            supa_truncate('alertas_meteo')  # vaciar tabla
            return

        supa_truncate('alertas_meteo')
        n = supa_post('alertas_meteo', unique[:50])
        print(f"   [OK] {n} alertas guardadas")
    except Exception as e:
        print(f"   [ERR] {e}")


# =====================================================================
# 4) PRESAS Y CUENCAS — JSON SINA
# =====================================================================
def scrape_presas_cuencas():
    """
    Endpoint JSON SINA. Si falla, scrape HTML.
    Schema:
      fecha_corte (date), presa (text), estado (text), capacidad_total_hm3 (num),
      almacenamiento_hm3 (num), pct_almacenamiento (num),
      region_hidrologica (text), fuente (text)
    """
    print("\n[PRESAS] Niveles de presas principales...")
    url_json = "https://sinav30.conagua.gob.mx:8080/Common/PresasPrincipales/?valores=true"
    rows = []
    hoy = dt.date.today().isoformat()

    try:
        r = requests.get(url_json, headers=HEADERS, timeout=TIMEOUT, verify=False)
        if r.ok:
            try:
                data = r.json()
            except Exception:
                data = []
            arr = data if isinstance(data, list) else data.get('presas', [])
            for p in arr:
                rows.append({
                    'fecha_corte':            hoy,
                    'presa':                  (p.get('nombreoficial') or p.get('nombre') or '')[:200],
                    'estado':                 (p.get('estado') or '')[:80],
                    'capacidad_total_hm3':    _num(p.get('capacidadNAMO') or p.get('capacidad')),
                    'almacenamiento_hm3':     _num(p.get('almacenaactual') or p.get('almacenamiento')),
                    'pct_almacenamiento':     _num(p.get('llenano') or p.get('porcentaje')),
                    'region_hidrologica':     (p.get('region') or '')[:80],
                    'fuente':                 'SINA CONAGUA',
                })
            if rows:
                supa_truncate('presas_cuencas')
                n = supa_post('presas_cuencas', rows)
                print(f"   [OK] {n} presas guardadas via JSON SINA")
                return
    except Exception as e:
        print(f"   [WARN] JSON SINA fallo: {e}")

    # Fallback HTML
    print("   Fallback HTML...")
    url_html = "https://sinaweb.conagua.gob.mx/portal/Reportes/SituacionPresas.aspx"
    try:
        r = requests.get(url_html, headers=HEADERS, timeout=TIMEOUT, verify=False)
        if not r.ok:
            print(f"   [WARN] HTTP {r.status_code}")
            return
        soup = BeautifulSoup(r.text, 'lxml')
        for table in soup.find_all('table'):
            for tr in table.find_all('tr')[1:]:
                cells = [td.get_text(' ', strip=True) for td in tr.find_all('td')]
                if len(cells) < 4:
                    continue
                # Heuristica: primera columna nombre presa
                presa = cells[0][:200]
                if not presa or len(presa) < 4:
                    continue
                rows.append({
                    'fecha_corte':           hoy,
                    'presa':                 presa,
                    'estado':                cells[1][:80] if len(cells) > 1 else '',
                    'capacidad_total_hm3':   _num(cells[2]) if len(cells) > 2 else None,
                    'almacenamiento_hm3':    _num(cells[3]) if len(cells) > 3 else None,
                    'pct_almacenamiento':    _num(cells[4]) if len(cells) > 4 else None,
                    'region_hidrologica':    '',
                    'fuente':                'SINA HTML',
                })
        if rows:
            supa_truncate('presas_cuencas')
            n = supa_post('presas_cuencas', rows)
            print(f"   [OK] {n} presas guardadas via HTML")
        else:
            print("   [WARN] Sin filas extraibles del HTML.")
    except Exception as e:
        print(f"   [ERR] {e}")


# =====================================================================
# MAIN
# =====================================================================
def main():
    print("Tequio CONAGUA Scraper v3 (HTML/CSV puros)")
    print(f"   Supabase: {SUPABASE_URL}")
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    scrape_clima_municipal()
    scrape_alertas_meteo()
    scrape_monitor_sequia()
    scrape_presas_cuencas()

    print("\nCONAGUA v3 scrape completo.")


if __name__ == '__main__':
    main()
